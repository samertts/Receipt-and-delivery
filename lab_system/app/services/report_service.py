import csv
from pathlib import Path

from lab_system.app.database import db as _db
from lab_system.app.settings.config import STORAGE_DIR

EXPORT_DIR = STORAGE_DIR / 'exports'


def _where_clause(date_from="", date_to=""):
    where = ["(r.deleted_at IS NULL OR r.deleted_at = '')"]
    params = []
    if date_from:
        where.append("r.created_at >= ?")
        params.append(f"{date_from}T00:00:00")
    if date_to:
        where.append("r.created_at <= ?")
        params.append(f"{date_to}T23:59:59")
    return " AND ".join(where), params


def receipt_summary(date_from="", date_to="", group_by="day"):
    clauses, params = _where_clause(date_from, date_to)
    with _db.get_conn() as conn:
        by_status = conn.execute(
            f"SELECT r.status, COUNT(*) cnt FROM receipts r WHERE {clauses} GROUP BY r.status",
            params,
        ).fetchall()
        by_type = conn.execute(
            f"""SELECT t.name tx_type, COUNT(*) cnt
                FROM receipts r JOIN transaction_types t ON t.id=r.tx_type_id
                WHERE {clauses} GROUP BY t.name""",
            params,
        ).fetchall()
        total = conn.execute(
            f"SELECT COUNT(*) c FROM receipts r WHERE {clauses}", params,
        ).fetchone()["c"]
    return {
        "total": total,
        "by_status": {r["status"]: r["cnt"] for r in by_status},
        "by_type": {r["tx_type"]: r["cnt"] for r in by_type},
    }


def daily_report(date_from="", date_to=""):
    clauses, params = _where_clause(date_from, date_to)
    with _db.get_conn() as conn:
        rows = conn.execute(
            f"""SELECT DATE(r.created_at) day,
                       COUNT(*) total,
                       SUM(CASE WHEN r.status='Approved' THEN 1 ELSE 0 END) approved,
                       SUM(CASE WHEN r.status='Rejected' THEN 1 ELSE 0 END) rejected
                FROM receipts r WHERE {clauses}
                GROUP BY day ORDER BY day DESC""",
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def monthly_report(year=None):
    clauses, params = _where_clause()
    year_filter = ""
    if year:
        year_filter = " AND SUBSTR(r.created_at,1,4)=?"
        params.append(str(year))
    with _db.get_conn() as conn:
        rows = conn.execute(
            f"""SELECT SUBSTR(r.created_at,1,7) month,
                       COUNT(*) total,
                       SUM(CASE WHEN r.status='Approved' THEN 1 ELSE 0 END) approved,
                       SUM(CASE WHEN r.status='Rejected' THEN 1 ELSE 0 END) rejected
                FROM receipts r WHERE {clauses} {year_filter}
                GROUP BY month ORDER BY month DESC""",
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def institution_statistics(date_from="", date_to=""):
    clauses, params = _where_clause(date_from, date_to)
    with _db.get_conn() as conn:
        senders = conn.execute(
            f"""SELECT so.name org_name, COUNT(*) cnt
                FROM receipts r JOIN organizations so ON so.id=r.sender_org_id
                WHERE {clauses} GROUP BY so.name ORDER BY cnt DESC""",
            params,
        ).fetchall()
        receivers = conn.execute(
            f"""SELECT ro.name org_name, COUNT(*) cnt
                FROM receipts r JOIN organizations ro ON ro.id=r.receiver_org_id
                WHERE {clauses} GROUP BY ro.name ORDER BY cnt DESC""",
            params,
        ).fetchall()
    return {
        "by_sender": [dict(r) for r in senders],
        "by_receiver": [dict(r) for r in receivers],
    }


def rejection_statistics(date_from="", date_to=""):
    clauses, params = _where_clause(date_from, date_to)
    with _db.get_conn() as conn:
        rows = conn.execute(
            f"""SELECT st.name sample_name,
                       SUM(ri.rejected_count) total_rejected,
                       SUM(ri.total_count) total_samples,
                       ROUND(CAST(SUM(ri.rejected_count) AS REAL) / NULLIF(SUM(ri.total_count),0) * 100, 1) rejection_pct
                FROM receipt_items ri
                JOIN receipts r ON r.id=ri.receipt_id
                JOIN sample_types st ON st.id=ri.sample_type_id
                WHERE {clauses}
                GROUP BY st.name ORDER BY total_rejected DESC""",
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def damage_statistics(date_from="", date_to=""):
    clauses, params = _where_clause(date_from, date_to)
    with _db.get_conn() as conn:
        rows = conn.execute(
            f"""SELECT st.name sample_name,
                       SUM(ri.damaged_count) total_damaged,
                       SUM(ri.total_count) total_samples,
                       ROUND(CAST(SUM(ri.damaged_count) AS REAL) / NULLIF(SUM(ri.total_count),0) * 100, 1) damage_pct
                FROM receipt_items ri
                JOIN receipts r ON r.id=ri.receipt_id
                JOIN sample_types st ON st.id=ri.sample_type_id
                WHERE {clauses}
                GROUP BY st.name ORDER BY total_damaged DESC""",
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def sample_summary(date_from="", date_to=""):
    clauses, params = _where_clause(date_from, date_to)
    with _db.get_conn() as conn:
        rows = conn.execute(
            f"""SELECT st.name sample_name,
                       SUM(ri.total_count) total,
                       SUM(ri.valid_count) valid,
                       SUM(ri.damaged_count) damaged,
                       SUM(ri.rejected_count) rejected,
                       SUM(ri.non_conforming_count) non_conf
                FROM receipt_items ri
                JOIN receipts r ON r.id=ri.receipt_id
                JOIN sample_types st ON st.id=ri.sample_type_id
                WHERE {clauses}
                GROUP BY st.name ORDER BY total DESC""",
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def export_receipts_csv(file_path, q="", status="", date_from="", date_to=""):
    from lab_system.app.services.receipt_service import list_receipts
    rows, _ = list_receipts(
        q=q, status=status, date_from=date_from, date_to=date_to, page=1, page_size=999999,
    )
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "رقم الإيصال", "النوع", "الجهة المرسلة", "الجهة المستقبلة",
            "اسم المرسل", "اسم المستلم", "التاريخ", "الحالة", "ملاحظات",
        ])
        for r in rows:
            writer.writerow([
                r["receipt_no"], r["tx_type"],
                r.get("sender_org", ""), r.get("receiver_org", ""),
                r.get("sender_name", ""), r.get("receiver_name", ""),
                r["created_at"], r["status"], r.get("notes", ""),
            ])
    return path


def export_xlsx(file_path, data_rows, headers):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    hdr_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def export_pdf(file_path, title, headers, data_rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 12))
    table_data = [headers] + [list(map(str, row)) for row in data_rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#D9E2F3")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return path
