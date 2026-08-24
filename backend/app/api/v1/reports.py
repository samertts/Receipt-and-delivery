from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.api.deps import require_permission
from app.core.response_envelope import wrap_response
from app.db.session import get_db
from app.models.organization import Organization
from app.models.transaction import Transaction
from app.models.user import User
from app.schemas.transaction import TransactionResponse

router = APIRouter(prefix="/reports", tags=["التقارير"])

_STATUS_LABELS = {
    "approved": "معتمدة",
    "draft": "مسودة",
    "rejected": "مرفوضة",
    "archived": "مؤرشفة",
    "cancelled": "ملغاة",
}
_PDF_FONT = "Helvetica"


def _start_datetime(value: date) -> datetime:
    return datetime.combine(value, time.min, tzinfo=timezone.utc)


def _end_datetime(value: date) -> datetime:
    return datetime.combine(value + timedelta(days=1), time.min, tzinfo=timezone.utc)


def _query_transactions(
    db: Session,
    start_date: date | None,
    end_date: date | None,
    status: str,
    transaction_type: str,
):
    query = db.query(Transaction).order_by(Transaction.created_at.desc())
    if start_date:
        query = query.filter(Transaction.created_at >= _start_datetime(start_date))
    if end_date:
        query = query.filter(Transaction.created_at < _end_datetime(end_date))
    if status:
        query = query.filter(Transaction.status == status)
    if transaction_type:
        query = query.filter(Transaction.transaction_type == transaction_type)
    return query


def _report_data(
    db: Session,
    start_date: date | None,
    end_date: date | None,
    status: str,
    transaction_type: str,
) -> dict[str, Any]:
    transactions = _query_transactions(
        db, start_date, end_date, status, transaction_type
    ).all()
    by_status: dict[str, int] = {}
    by_type: dict[str, int] = {}
    for transaction in transactions:
        by_status[transaction.status] = by_status.get(transaction.status, 0) + 1
        key = transaction.transaction_type or "غير محدد"
        by_type[key] = by_type.get(key, 0) + 1

    return {
        "filters": {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "status": status or None,
            "transaction_type": transaction_type or None,
        },
        "summary": {
            "total": len(transactions),
            "by_status": by_status,
            "by_type": by_type,
        },
        "transactions": transactions,
    }


def _transaction_rows(transactions: list[Transaction]) -> list[list[str]]:
    return [
        [
            str(transaction.transaction_no or ""),
            str(transaction.transaction_type or ""),
            str(transaction.sender_name or ""),
            str(transaction.receiver_name or ""),
            _STATUS_LABELS.get(transaction.status, transaction.status or ""),
            transaction.transaction_date or "",
            transaction.created_at.isoformat() if transaction.created_at else "",
        ]
        for transaction in transactions
    ]


def _safe_filename_header(filename: str) -> dict[str, str]:
    return {"Content-Disposition": f'attachment; filename="{filename}"'}


@router.get("/summary")
def get_report_summary(
    start_date: date | None = Query(None, description="تاريخ البداية"),
    end_date: date | None = Query(None, description="تاريخ النهاية"),
    status: str = Query("", description="الحالة"),
    transaction_type: str = Query("", description="نوع المعاملة"),
    current_user: User = Depends(require_permission("view_reports")),
    db: Session = Depends(get_db),
):
    if start_date and end_date and start_date > end_date:
        from app.core.exceptions import ValidationError

        raise ValidationError("تاريخ البداية يجب أن يسبق تاريخ النهاية")

    report = _report_data(db, start_date, end_date, status, transaction_type)
    report["summary"]["by_status"] = {
        key: report["summary"]["by_status"].get(key, 0) for key in _STATUS_LABELS
    } | {
        key: value
        for key, value in report["summary"]["by_status"].items()
        if key not in _STATUS_LABELS
    }
    transactions = report.pop("transactions")
    report["transactions"] = [
        TransactionResponse.model_validate(transaction).model_dump(mode="json")
        for transaction in transactions
    ]
    return wrap_response(data=report, message="تم تحميل التقرير بنجاح")


@router.get("/transactions.xlsx")
def export_transactions_excel(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    status: str = Query(""),
    transaction_type: str = Query(""),
    _: User = Depends(require_permission("view_reports")),
    db: Session = Depends(get_db),
):
    report = _report_data(db, start_date, end_date, status, transaction_type)
    workbook = Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "المعاملات"
    transactions_sheet.sheet_view.rightToLeft = True
    headers = [
        "رقم المعاملة",
        "نوع المعاملة",
        "المرسل",
        "المستلم",
        "الحالة",
        "تاريخ المعاملة",
        "تاريخ الإنشاء",
    ]
    transactions_sheet.append(["تقرير المعاملات"])
    transactions_sheet.append([f"الفلاتر: {report['filters']}"])
    transactions_sheet.append(headers)
    for row in _transaction_rows(report["transactions"]):
        transactions_sheet.append(row)
    transactions_sheet.freeze_panes = "A4"
    transactions_sheet.auto_filter.ref = f"A3:G{max(3, transactions_sheet.max_row)}"

    stats_sheet = workbook.create_sheet("الإحصائيات")
    stats_sheet.sheet_view.rightToLeft = True
    stats_sheet.append(["ملخص الإحصائيات"])
    stats_sheet.append(["المؤشر", "القيمة"])
    stats_sheet.append(["إجمالي المعاملات", report["summary"]["total"]])
    stats_sheet.append([])
    stats_sheet.append(["الحالة", "العدد"])
    for key in _STATUS_LABELS:
        stats_sheet.append([_STATUS_LABELS[key], report["summary"]["by_status"].get(key, 0)])
    stats_sheet.append([])
    stats_sheet.append(["نوع المعاملة", "العدد"])
    for key, count in sorted(
        report["summary"]["by_type"].items(), key=lambda item: item[1], reverse=True
    ):
        stats_sheet.append([key, count])

    for sheet in (transactions_sheet, stats_sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")
        for cell in sheet[1]:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3A5F")
        for cell in sheet[3] if sheet is transactions_sheet else sheet[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 38)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_safe_filename_header("transactions_report.xlsx"),
    )


def _register_pdf_font() -> str:
    global _PDF_FONT
    font_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", str(font_path)))
            _PDF_FONT = "DejaVuSans"
        except Exception:
            pass
    return _PDF_FONT


def _pdf_value(value: Any) -> str:
    text = str(value or "")
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        return get_display(arabic_reshaper.reshape(text))
    except ImportError:
        return text


@router.get("/transactions.pdf")
def export_transactions_pdf(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    status: str = Query(""),
    transaction_type: str = Query(""),
    _: User = Depends(require_permission("view_reports")),
    db: Session = Depends(get_db),
):
    report = _report_data(db, start_date, end_date, status, transaction_type)
    font_name = _register_pdf_font()
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName=font_name,
        alignment=TA_CENTER,
        fontSize=16,
        leading=20,
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontName=font_name,
        alignment=TA_RIGHT,
        fontSize=8,
        leading=11,
    )
    story = [
        Paragraph(_pdf_value("تقرير المعاملات والإحصائيات"), title_style),
        Spacer(1, 5 * mm),
        Paragraph(
            _pdf_value(
                f"إجمالي المعاملات: {report['summary']['total']} | "
                f"الفترة: {report['filters']['start_date'] or 'الكل'} - "
                f"{report['filters']['end_date'] or 'الكل'}"
            ),
            body_style,
        ),
        Spacer(1, 4 * mm),
    ]
    headers = ["رقم المعاملة", "النوع", "المرسل", "المستلم", "الحالة", "التاريخ"]
    table_data = [[Paragraph(_pdf_value(header), body_style) for header in headers]]
    for row in _transaction_rows(report["transactions"]):
        table_data.append(
            [Paragraph(_pdf_value(value), body_style) for value in [row[0], row[1], row[2], row[3], row[4], row[5]]]
        )
    if len(table_data) == 1:
        table_data.append([Paragraph(_pdf_value("لا توجد بيانات"), body_style)] + [""] * 5)
    table = Table(table_data, repeatRows=1, colWidths=[34 * mm, 34 * mm, 48 * mm, 48 * mm, 28 * mm, 32 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(table)
    document.build(story)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=_safe_filename_header("transactions_report.pdf"),
    )


@router.get("/governorate/{governorate_name}")
async def get_governorate_report(
    governorate_name: str,
    start_date: date = Query(..., description="تاريخ البداية"),
    end_date: date = Query(..., description="تاريخ النهاية"),
    current_user: User = Depends(require_permission("view_reports")),
    db: Session = Depends(get_db),
):
    """Aggregate data across all institutions in a governorate."""
    institutions = (
        db.query(Organization)
        .filter(Organization.governorate == governorate_name)
        .all()
    )
    if not institutions:
        return {
            "governorate": governorate_name,
            "total_transactions": 0,
            "institutions": [],
        }

    institution_ids = [str(inst.id) for inst in institutions]
    transaction_counts = (
        db.query(Transaction.sender_organization_id, func.count(Transaction.id))
        .filter(
            Transaction.sender_organization_id.in_(institution_ids),
            Transaction.created_at >= _start_datetime(start_date),
            Transaction.created_at < _end_datetime(end_date),
        )
        .group_by(Transaction.sender_organization_id)
        .all()
    )
    count_map = dict(transaction_counts)
    institutions_data = []
    total_transactions = 0
    for inst in institutions:
        count = count_map.get(str(inst.id), 0)
        total_transactions += count
        institutions_data.append(
            {
                "id": str(inst.id),
                "name": inst.name,
                "code": inst.code,
                "transaction_count": count,
            }
        )
    return {
        "governorate": governorate_name,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_transactions": total_transactions,
        "institutions": institutions_data,
    }
