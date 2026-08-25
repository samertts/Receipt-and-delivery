from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader


def test_report_summary_returns_filtered_report_contract(client, admin_token):
    response = client.get(
        '/api/reports/summary',
        params={'status': 'approved'},
        headers={'Authorization': f'Bearer {admin_token}'},
    )

    assert response.status_code == 200
    body = response.json()
    assert body['success'] is True
    assert body['data']['filters']['status'] == 'approved'
    assert body['data']['summary']['total'] == 0
    assert body['data']['transactions'] == []


def test_english_report_summary_and_excel_use_english_labels(client, admin_token):
    summary = client.get(
        '/api/reports/summary',
        params={'lang': 'en'},
        headers={'Authorization': f'Bearer {admin_token}'},
    )
    assert summary.status_code == 200
    assert summary.json()['message'] == 'Report loaded successfully'

    response = client.get(
        '/api/reports/transactions.xlsx',
        params={'lang': 'en'},
        headers={'Authorization': f'Bearer {admin_token}'},
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ['Transactions', 'Statistics']
    assert workbook['Transactions']['A1'].value == 'Transaction Report'
    assert workbook['Transactions']['A3'].value == 'Transaction number'
    assert workbook['Statistics']['A2'].value == 'Metric'


def test_excel_export_returns_a_readable_workbook(client, admin_token):
    response = client.get(
        '/api/reports/transactions.xlsx',
        headers={'Authorization': f'Bearer {admin_token}'},
    )

    assert response.status_code == 200
    assert response.headers['content-type'].startswith(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ['المعاملات', 'الإحصائيات']


def test_pdf_export_returns_a_readable_pdf(client, admin_token):
    response = client.get(
        '/api/reports/transactions.pdf',
        headers={'Authorization': f'Bearer {admin_token}'},
    )

    assert response.status_code == 200
    assert response.headers['content-type'] == 'application/pdf'
    reader = PdfReader(BytesIO(response.content))
    assert len(reader.pages) == 1
